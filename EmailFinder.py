# -*- coding: utf-8 -*-
"""
Created on Sun May 08 00:47:47 2016

@author: op3ntrap@gmail.com

Email finder within an excel sheet
uses string validations to qualify emails.
uses multithreading for faster execution.
Returns an array with a non duplicate list
"""
from openpyxl import load_workbook
from threading import Thread
from flanker.addresslib import address
from openpyxl import Workbook


result=[]
input_file=""
def find_email_cells(input_file,save):
    file_name = input_file    
    current_workbook = load_workbook(filename=file_name)
    Sheets= current_workbook.get_sheet_names()
    def check_for_emails(active_sheet):
      
        for sample_row in active_sheet.rows:
            for sample_cell in sample_row:
                try:
                    if "@" in sample_cell.value:
                        if "." in sample_cell.value:
                            if address.parse(sample_cell.value,addr_spec_only=True) == sample_cell.value:
                                if sample_cell.value not in result:
                                    result.append(sample_cell.value)
                                    print sample_cell.value
                except TypeError:
                    continue
     
    #find cells with emails and add it to the results array
    local_threads = []            
    for sheet in Sheets:
        actie_sheet = current_workbook.get_sheet_by_name(sheet)    
        sheet_thread = Thread(target=check_for_emails,args =(actie_sheet,))
        sheet_thread.start()
        local_threads.append(sheet_thread)
        
    for t in local_threads:
        t.join()
    if save == True:
        save_file()
        
    return result
    
    
    
  
#print result
def save_file():
    result_workbook = Workbook()
    ws1 = result_workbook.create_sheet(title="Emails")
    ws1.cell(row=1,column=1).value = "Email"
    j=2
    for email in result:
        ws1.cell(row=j,column=1).value = email
        j=j+1
    #path_input = ""#user input path
    final_path = "Emails_" + input_file
    result_workbook.save(final_path)






            
    
        
    
        
    